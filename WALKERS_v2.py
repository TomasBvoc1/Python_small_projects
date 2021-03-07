import openpyxl as op
import pandas as pd
import csv

class Walkers:
    def __init__(self, key, team_no, team_name, email, name, initial_steps):
        self.key = key
        self.team_no = team_no
        self.team_name = team_name
        self.email = email
        self.name = name
        self.steps = initial_steps
        #self.transaction = False

class Teams:
    def __init__(self, key, team_no, team_name, initial_steps):
        self.key = key
        self.team_no = team_no
        self.team_name = team_name
        self.steps = initial_steps

try:
    walkers_df = pd.read_excel('/Users/TomasBvoc/Box Sync/WALKERS/WALKERS_assignement.xlsx', sheet_name='initial')
    walkers_list = []
    teams_df = pd.read_excel('/Users/TomasBvoc/Box Sync/WALKERS/WALKERS_assignement.xlsx', sheet_name='teams')
    teams_list = []
except Exception as e:
    print(e)
    sys.exit(1)

# Standard syntax for iterating a dataframe
for index, row in walkers_df.iterrows():
    key = row['Key']
    team_no = row['Team No.']
    team_name = row['Team Name']
    email = row['email']
    name = row['Name']
    initial_steps = row['Initial steps']
    temp_object = Walkers(key, team_no, team_name, email, name, initial_steps)
    walkers_list.append(temp_object)

for index, row in teams_df.iterrows():
    key = row['Key']
    team_no = row['Team No.']
    team_name = row['Team Name']
    team_initial_steps = row['Initial steps']
    temp_object = Teams(key, team_no, team_name, team_initial_steps)
    teams_list.append(temp_object)

print("We can print list of Teams this way...\n",teams_df)  #OK!       yeah
print("...or we can print list of Teams this way: ")        #OK too !  happy
for item in teams_list:
    print(item.team_name)

print("-----------------------------------------------------------------------------------------------------------------------------")
print("The Walk-it-out activity has",len(walkers_list), "members in 6 teams, and the initial steps are (zero for everyone naturally) :")
for item in walkers_list:
    print(item.team_name, item.name, item.steps)

def find_team(team_id):
    for item in teams_list:
        if item.team_name == team_id:
            return item

steps_df = pd.read_csv('/Users/TomasBvoc/Box Sync/WALKERS/CSV_input_from_Publisher/Week3_February_1to7/WalkItOutweek3_endofweek.csv')
for index, column in steps_df.iterrows():
    temp_team_name = column[3]
    new_steps = column[5]
    temp_object = find_team(temp_team_name)
    temp_object.steps = temp_object.steps + int(new_steps)
    #temp_object.transaction = True

print("\nCurrent steps after Week0 are :")
for item in teams_list:
    print(item.team_name, item.steps)

#-----------------------------------------------------------------
print("\nNow, We're gonna try to save results into Excel file 'WALKERS_??????.xlsx'... ")
wb = op.Workbook()
ws = wb.active
ws.title = "subtotal_steps"

ws.cell(row=1, column=1, value="Team")
ws.cell(row=1, column=2, value="Steps Week 1")
ws.cell(row=1, column=3, value="Steps Week 2")
ws.cell(row=1, column=4, value="Steps Week 3")
ws.cell(row=1, column=5, value="Steps Week 4")
ws.cell(row=1, column=6, value="Steps TOTAL")

row_number = 2
for item in teams_list:
    ws.cell(row=row_number, column=1, value=item.team_name)
    ws.cell(row=row_number, column=4, value=item.steps)
    row_number += 1

ws.column_dimensions['A'].width = 30
columns = ['B','C','D','E','F']
for item in columns:
    ws.column_dimensions[item].width = 15

wb.save("/Users/TomasBvoc/Box Sync/WALKERS/XLS_output_from_Pycharm/WALKERS_after_week3.xlsx")
wb.close()